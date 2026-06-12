import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


# -----------------------------
# BASE PATH (CONTAINER READY)
# -----------------------------
BASE_DIR = Path(os.getenv("APP_BASE_DIR", Path.cwd()))

DOWNLOAD_DIR = BASE_DIR / "downloads"
PROCESSED_DIR = BASE_DIR / "processed"

DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)


# -----------------------------
# FILE LOADER (MULTI FORMAT)
# -----------------------------
def read_file(file_path: str, has_header: bool = True):
    file_path = Path(file_path)

    if file_path.suffix == ".xlsx":
        return read_excel_flexible(file_path, has_header)

    elif file_path.suffix == ".csv":
        return read_csv_flexible(file_path, has_header)

    else:
        raise ValueError(f"Unsupported file type: {file_path.suffix}")


# -----------------------------
# CSV HANDLER
# -----------------------------
def read_csv_flexible(file_path: str, has_header: bool = True):
    if has_header:
        df = pd.read_csv(file_path, dtype=str)
    else:
        df = pd.read_csv(file_path, header=None, dtype=str)
        df.columns = [f"Column_{i+1}" for i in range(len(df.columns))]

    df = df.fillna("")
    df.columns = [str(col).strip() for col in df.columns]

    return df


# -----------------------------
# EXCEL TABLE (PANDAS)
# -----------------------------
def read_excel_flexible(file_path: str, has_header: bool = True):
    sheet = 0  # første ark

    if has_header:
        df = pd.read_excel(file_path, dtype=str, engine="openpyxl", sheet_name=sheet)
    else:
        df = pd.read_excel(file_path, header=None, dtype=str, engine="openpyxl", sheet_name=sheet)
        df.columns = [f"Column_{i+1}" for i in range(len(df.columns))]

    df = df.fillna("")
    df.columns = [str(col).strip() for col in df.columns]

    return df


# -----------------------------
# EXCEL CELL HANDLER
# -----------------------------
def excel_edit(
    file_path: str,
    mode: str,
    from_ref: str = None,
    to_ref: str = None,
    value=None,
    sheet_name: str = None,
    save: bool = False,
    save_path: str = None
):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    print(f"Bruger ark: '{ws.title}'")

    # ---------- READ ----------
    if mode == "read":
        if not from_ref:
            raise ValueError("read kræver from_ref")

        if ":" in from_ref:
            min_col, min_row, max_col, max_row = range_boundaries(from_ref)

            data = []
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col
            ):
                data.append([cell.value for cell in row])

            return data

        return ws[from_ref].value

    # ---------- COPY ----------
    elif mode == "copy":
        if not from_ref or not to_ref:
            raise ValueError("copy kræver både from_ref og to_ref")

        if ":" in from_ref and ":" in to_ref:
            min_c1, min_r1, max_c1, max_r1 = range_boundaries(from_ref)
            min_c2, min_r2, max_c2, max_r2 = range_boundaries(to_ref)

            # ✅ validation
            rows1 = max_r1 - min_r1 + 1
            cols1 = max_c1 - min_c1 + 1
            rows2 = max_r2 - min_r2 + 1
            cols2 = max_c2 - min_c2 + 1

            if rows1 != rows2 or cols1 != cols2:
                raise ValueError("Ranges skal have samme størrelse")

            for i, row in enumerate(ws.iter_rows(
                min_row=min_r1,
                max_row=max_r1,
                min_col=min_c1,
                max_col=max_c1
            )):
                for j, cell in enumerate(row):
                    ws.cell(row=min_r2 + i, column=min_c2 + j).value = cell.value
        else:
            ws[to_ref] = ws[from_ref].value

    # ---------- WRITE ----------
    elif mode == "write":
        if not to_ref:
            raise ValueError("write kræver to_ref")

        if ":" in to_ref:
            min_col, min_row, max_col, max_row = range_boundaries(to_ref)

            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col
            ):
                for cell in row:
                    cell.value = value
        else:
            ws[to_ref] = value

    # ---------- CLEAR ----------
    elif mode == "clear":
        if not to_ref:
            raise ValueError("clear kræver to_ref")

        if ":" in to_ref:
            min_col, min_row, max_col, max_row = range_boundaries(to_ref)

            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col
            ):
                for cell in row:
                    cell.value = None
        else:
            ws[to_ref] = None

    else:
        raise ValueError("Invalid mode")

    # ---------- SAVE ----------
    if save:
        if not save_path:
            raise ValueError("save_path skal angives når save=True")

        save_path = Path(save_path)

        # ✅ container-safe (relative paths → processed)
        if not save_path.is_absolute():
            save_path = PROCESSED_DIR / save_path

        save_path.parent.mkdir(parents=True, exist_ok=True)

        print(f"Gemmer fil til: {save_path}")

        wb.save(save_path)

    return


# -----------------------------
# SAFE WRAPPER
# -----------------------------
def safe_edit_file(
    file_path: str,
    mode: str = None,
    from_ref: str = None,
    to_ref: str = None,
    value=None,
    sheet_name: str = None,
    save: bool = False,
    save_path: str = None,
    has_header: bool = True
):
    file_path = Path(file_path)

    # ✅ container-safe input
    if not file_path.is_absolute():
        file_path = DOWNLOAD_DIR / file_path

    if file_path.suffix == ".csv":
        df = read_csv_flexible(file_path, has_header)

        if mode == "read":
            return df

        raise ValueError("CSV understøtter ikke cell manipulation")

    elif file_path.suffix == ".xlsx":

        if mode == "read":
            return read_excel_flexible(file_path, has_header)

        return excel_edit(
            file_path=str(file_path),
            mode=mode,
            from_ref=from_ref,
            to_ref=to_ref,
            value=value,
            sheet_name=sheet_name,
            save=save,
            save_path=save_path
        )

    else:
        raise ValueError(f"Unsupported file type: {file_path.suffix}")